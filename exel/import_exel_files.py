from typing import Type

import openpyxl
import pyexcel as p
import os

from models.parameter_config import ParameterConfig
from models.colum_property import ColumProperty
from models.table_property import TableProperty
from models.foreign_key import ForeignKey
from models.index_property import IndexProperty


class ImportExcelFile:
    """
    エクセルからデータベース作成用ファイルを作成する。
    """
    config = ParameterConfig
    exel_info = ''
    table_info_list = []

    def __init__(self, g: Type[ParameterConfig]):
        """
        初期化
        :param g:
        """
        self.config = g
        pass

    def import_exel(self):
        """
        エクセルを取り込みデータベースのオブジェクトに変換する
       :return:
        """
        # 作業用のxlsxファイルの設定(xlsファイルは読み込めないため)
        exel_dir_path = os.path.split(self.config.input_files_path)
        job_exel_file_path = exel_dir_path[0] + '/import_job.xlsx'

        # 作業用にxlsxファイルを保存する
        p.save_book_as(file_name=self.config.input_files_path,
                       dest_file_name=job_exel_file_path)

        # エクセル情報の取り込み
        self.exel_info = openpyxl.load_workbook(job_exel_file_path)

        # テーブル情報を作成する
        self.__make_table_info_list()

        # 外部キー情報を設定する
        self.__make_foreign_keys()

        # インデックス情報を設定する
        self.__make_index_list()

        # 作業で使用したエクセルは削除する
        os.remove(job_exel_file_path)

    def __make_table_info_list(self):
        """
        エクセルからデータベースの情報を作成する
        :return:
        """
        # 「全ての属性」からテーブル情報を取得する
        sheet = self.exel_info['全ての属性']
        print('最大列: ' + str(sheet.max_column))
        print('最大行: ' + str(sheet.max_row))

        # 一時保存の変数
        tmp_is_created_at_flag = False

        # テーブル情報を初期化する
        table_info = TableProperty()

        # すべてのセルを読み込む(ヘッダー部分は読み込まないため2から開始する)
        for row_num in range(2, sheet.max_row + 1):

            # エクセルから読み取った情報
            table_name_display = sheet['A' + str(row_num)].value  # テーブル名 (論理名)
            table_name = sheet['B' + str(row_num)].value  # テーブル名 (物理名)
            colum_name_display = sheet['C' + str(row_num)].value  # カラム名 (論理名)
            colum_name = sheet['D' + str(row_num)].value  # カラム名 (物理名)
            colum_type = sheet['E' + str(row_num)].value  # 型
            colum_length = sheet['F' + str(row_num)].value  # 長さ
            colum_decimal = sheet['G' + str(row_num)].value  # 小数
            colum_pk = sheet['H' + str(row_num)].value  # PK
            colum_not_null = sheet['I' + str(row_num)].value  # NOT NULL
            colum_unique = sheet['J' + str(row_num)].value  # UNIQUE
            colum_fk = sheet['K' + str(row_num)].value  # FK
            colum_auto_increment = sheet['L' + str(row_num)].value  # オートインクリメント
            colum_default_value = sheet['M' + str(row_num)].value  # デフォルト値
            colum_description = sheet['N' + str(row_num)].value  # 説明

            # エクセルのテーブル名が変わったらテーブル情報を新規に初期化する
            if table_name != table_info.table_name:
                print('------------' + table_name + '(' + table_name_display + ')----------------')

                # テーブル名がまだ設定されていない場合はテーブル情報は追加しない
                if table_info.table_name != '':
                    # テーブル情報を追加する
                    self.table_info_list.append(table_info)

                # データの初期化
                table_info = TableProperty()
                table_info.table_name = table_name
                table_info.table_display_name = table_name_display
                table_info.date_format = self.config.date_format
                tmp_is_created_at_flag = False

            # カラム情報の初期化を行う
            colum_info = ColumProperty()
            colum_info.colum_name = colum_name
            colum_info.colum_comment = colum_name_display
            colum_info.colum_type = colum_type
            colum_info.database_colum_type = colum_type
            colum_info.length = colum_length
            colum_info.decimal = colum_decimal
            colum_info.default_value = colum_default_value
            colum_info.description = colum_description

            # timestamps フラグかどうかを設定する
            if tmp_is_created_at_flag \
                    and (colum_info.colum_name == 'created_at' or colum_info.colum_name == 'updated_at'):
                table_info.is_timestamps = True

            # created フラグを設定する
            if not tmp_is_created_at_flag \
                    and (colum_info.colum_name == 'created_at' or colum_info.colum_name == 'updated_at'):
                tmp_is_created_at_flag = True

            # 主キーが存在するかどうかを確認して、主キーであれば主キーの配列を追加する
            if colum_pk is not None and colum_pk.find('○') != -1:
                table_info.primary_keys.append(colum_name)
                colum_info.is_primary = True

            # このカラムがタイムスタンプの場合には「dates」に追加する
            if colum_info.database_colum_type.lower().find('timestamp') != -1:
                table_info.dates.append(colum_info.colum_name)

            # 「deleted_at」が存在するテーブルはソフトデリート対応とする
            if colum_info.colum_name == 'deleted_at':
                table_info.is_soft_deleted = True

            # オートインクリメントが存在するか確認して、存在した場合はテーブルとカラムの両方のインクリメント情報をTrueにする
            # またはテーブルの設定が「serial」または「bigserial」の場合も、テーブルとカラムの両方のインクリメントをTrueにする
            if (colum_auto_increment is not None and colum_auto_increment == '○') or \
                    (colum_info.colum_name.lower().find('serial') != -1 or
                     colum_info.colum_name.lower().find('bigserial') != -1):
                table_info.is_incrementing = True
                colum_info.is_auto_increment = True

            # NotNullかどうかを判定して、存在した場合はカラム情報のNotNullをTrueにする
            if colum_not_null is not None and colum_not_null.find('○') != -1:
                colum_info.is_not_null = True

            # Uniqueかどうかを判定して、存在した場合はカラム情報のUniqueをTrueにする
            if colum_unique is not None and colum_unique.find('○') != -1:
                colum_info.is_unique = True

            # カラム情報を追加する
            table_info.column_list.append(colum_info)
            print(' - ' + colum_name + '(' + colum_name_display + ')')

        # ループの最後の行はデータが追加できないため、このタイミングでテーブル情報を追加する
        self.table_info_list.append(table_info)

    def __make_foreign_keys(self):
        """
        紐付け情報を設定する
        :return:
        """
        sheet = self.exel_info['全てのテーブル']

        # ループ中に使用する一時的なフラグ
        is_foreign_key_flag = False     # 外部キーフラグ
        tmp_foreign_key_list = []       # 一時的な外部キーリスト
        table_name = ''

        # すべてのセルを読み込む
        for row_num in range(1, sheet.max_row + 1):
            cell_name = sheet['A' + str(row_num)].value  # Aのセル名称

            # 外部キー設定モードであれば外部キー情報を設定する
            if is_foreign_key_flag \
                    and cell_name is not None \
                    and cell_name != 'カラム名 (論理名)':
                foreign_key = ForeignKey()
                foreign_key.colum_name_display = sheet['A' + str(row_num)].value  # カラム名 (論理名)
                foreign_key.colum_name = sheet['B' + str(row_num)].value  # カラム名 (物理名)
                foreign_key.ref_table = sheet['C' + str(row_num)].value  # 参照テーブル
                foreign_key.ref_key = sheet['D' + str(row_num)].value  # 参照キー
                tmp_foreign_key_list.append(foreign_key)
                continue

            # 外部キー設定モードで空白になればモードを解除する
            if is_foreign_key_flag and cell_name is None:
                is_foreign_key_flag = False     # 外部キーフラグをFalseにする

            # セル名が「テーブル名 (物理名)」であれば新しいテーブル情報と判断して新規に
            if cell_name == 'テーブル名 (物理名)' or cell_name == 'Table name (physical name)':

                # 該当するテーブル情報に外部キー情報を取得する
                if table_name != '':
                    for table_info in self.table_info_list:
                        # 該当するテーブル名のテーブル情報に対して外部キー情報を設定する
                        if table_info.table_name == table_name:
                            table_info.foreign_keys = tmp_foreign_key_list
                            break   # 情報を設定すると、これ以上の処理は無駄なので抜ける

                # 初期化処理
                table_name = sheet['B' + str(row_num)].value  # テーブル名を設定する
                tmp_foreign_key_list = []

            elif cell_name == '外部キー' or cell_name == 'FOREIGN KEY':
                # 「外部キー」のタイトルを見つけると次のセルから外部キーの情報が記載されているので、外部キー設定のフラグをTrueにする
                is_foreign_key_flag = True

    def __make_index_list(self):
        """
        インデック情報一覧を設定する
        :return:
        """
        sheet = self.exel_info['全てのインデックス']

        is_index_colum_mode = False     # インデックスのカラム追加状態
        index_property = None           # インデックス情報を初期化
        tmp_index_property_list = []    # インデックス情報の一覧

        # すべてのセルを読み込む
        for row_num in range(1, sheet.max_row + 1):
            cell_name = sheet['A' + str(row_num)].value  # Aのセル名称

            # インデックスカラムモードとなっており、空白となった場合にインデックスカラムモードを解除する
            if is_index_colum_mode and (cell_name is None or cell_name == ''):
                is_index_colum_mode = False
                continue

            # インデックスカラムモードで値があれば対象となるインデックスカラムを追加する
            if is_index_colum_mode and (cell_name is not None and cell_name != ''):
                index_property.colum_list.append(sheet['C' + str(row_num)].value)
                continue

            # インデック名を取得する
            if cell_name == 'インデックス名':
                is_index_colum_mode = False     # インデックスのカラム取得状態にする

                # インデックス情報が存在していた場合は処理を追加する
                if index_property is not None:
                    tmp_index_property_list.append(index_property)

                # インデックス情報の初期化
                index_property = IndexProperty()
                index_property.index_name = sheet['B' + str(row_num)].value  # インデックス名を取得する
                continue

            # タイプを取得する
            if cell_name == 'タイプ':
                index_property.type = sheet['B' + str(row_num)].value  # インデックス名を取得する
                continue

            # ユニーク
            if cell_name == 'UNIQUE':
                if sheet['B' + str(row_num)].value == '○':
                    index_property.is_unique = True
                continue

            # 説明
            if cell_name == '説明':
                index_property.description = sheet['B' + str(row_num)].value  # 説明を取得する
                index_property.table_name = sheet['B' + str(row_num + 1)].value  # 対象となるテーブル名は次の行に設定されている
                continue

            # インデックスカラムの設定
            if cell_name == 'カラムの順番':
                is_index_colum_mode = True  # カラムの順番があればインデックスカラムの設定
                continue

        # 最後に残っていた場合はデータを追加する
        if index_property is not None:
            tmp_index_property_list.append(index_property)

        # インデックス情報を該当するテーブルのインデックス情報に追加する
        for table_info in self.table_info_list:
            for index_property in tmp_index_property_list:
                if table_info.table_name == index_property.table_name:
                    table_info.index_list.append(index_property)
